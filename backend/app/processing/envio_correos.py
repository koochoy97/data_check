"""
Bloque 2: Envío de Correos (Personas + Empresas + Reply Rates)
Adaptado de auditoría_elt.py (líneas 279-612)

Input: CSV con columnas "Delivery date", "Contacted", "Replied", "Delivered", "Contact email"
Output: Agrega a un spreadsheet existente:
  - Hoja "Base Correos" con datos + columnas derivadas
  - 8 pivots: Personas/Empresas × Día/Semana/Mes/Quarter
  - 4 hojas Reply Rates: Diario/Semanal/Mensual/Quarterly
"""
import pandas as pd
from openpyxl.utils import get_column_letter


def procesar_correos(csv_path, spreadsheet, service):
    """
    Args:
        csv_path: Path al CSV de Email Activity (contact-specific)
        spreadsheet: gspread Spreadsheet object (ya creado)
        service: Google Sheets API service
    Returns:
        dict con metadata del procesamiento
    """
    SPREADSHEET_ID = spreadsheet.id

    # ── Leer CSV ──
    df = pd.read_csv(csv_path)

    # ── Columna derivada: Delivery Date Text ──
    parsed_dates = pd.to_datetime(
        df["Delivery date"],
        format="%a, %d %b %Y %H:%M:%S",
        errors="coerce",
    )

    df["Delivery Date Text"] = parsed_dates.where(
        df["Contacted"].astype(str) == "1"
    ).dt.strftime("%d/%m/%Y")

    # Keep original df for reply rate calculations
    df_original = df.copy()

    df = df.replace([float("inf"), float("-inf")], "")
    df = df.fillna("")
    df = df.astype(str)

    # ── Crear hoja "Base Correos" ──
    sheet = spreadsheet.add_worksheet(
        title="Base Correos",
        rows=str(len(df) + 10),
        cols=str(len(df.columns) + 10),
    )
    sheet.update([df.columns.tolist()] + df.values.tolist())

    BASE_SHEET_ID = sheet._properties["sheetId"]
    last_row = len(df) + 1

    # ── Índices ──
    TEXT_COL_INDEX = df.columns.get_loc("Delivery Date Text")
    DATE_COL_INDEX = TEXT_COL_INDEX + 1
    WEEK_COL_INDEX = DATE_COL_INDEX + 1
    MONTH_COL_INDEX = DATE_COL_INDEX + 2
    QUARTER_COL_INDEX = DATE_COL_INDEX + 3
    EMAIL_COL_INDEX = df.columns.get_loc("Contact email")
    DOMAIN_COL_INDEX = QUARTER_COL_INDEX + 1

    col_letter = get_column_letter(TEXT_COL_INDEX + 1)
    date_letter = get_column_letter(DATE_COL_INDEX + 1)
    email_letter = get_column_letter(EMAIL_COL_INDEX + 1)

    # ── Expandir grid ──
    service.spreadsheets().batchUpdate(
        spreadsheetId=SPREADSHEET_ID,
        body={"requests": [{"updateSheetProperties": {
            "properties": {"sheetId": BASE_SHEET_ID, "gridProperties": {"columnCount": DOMAIN_COL_INDEX + 1}},
            "fields": "gridProperties.columnCount",
        }}]},
    ).execute()

    # ── Columnas derivadas ──
    requests = [
        _header_cell(BASE_SHEET_ID, DATE_COL_INDEX, "Delivery Date (Date)"),
        _formula_cell(BASE_SHEET_ID, DATE_COL_INDEX,
                      f'=IF({col_letter}2="","",DATE(RIGHT({col_letter}2,4),MID({col_letter}2,4,2),LEFT({col_letter}2,2)))'),
        _header_cell(BASE_SHEET_ID, WEEK_COL_INDEX, "Semana del año"),
        _formula_cell(BASE_SHEET_ID, WEEK_COL_INDEX,
                      f'=IF({date_letter}2="","",TEXT(WEEKNUM({date_letter}2,21),"00")&" - "&TEXT({date_letter}2,"YY"))'),
        _header_cell(BASE_SHEET_ID, MONTH_COL_INDEX, "Mes - Año"),
        _formula_cell(BASE_SHEET_ID, MONTH_COL_INDEX,
                      f'=IF({date_letter}2="","",TEXT({date_letter}2,"MM")&" - "&TEXT({date_letter}2,"YY"))'),
        _header_cell(BASE_SHEET_ID, QUARTER_COL_INDEX, "Quarter - Año"),
        _formula_cell(BASE_SHEET_ID, QUARTER_COL_INDEX,
                      f'=IF({date_letter}2="","","Q"&ROUNDUP(MONTH({date_letter}2)/3,0)&" - "&TEXT({date_letter}2,"YY"))'),
        _header_cell(BASE_SHEET_ID, DOMAIN_COL_INDEX, "Email Domain"),
        _formula_cell(BASE_SHEET_ID, DOMAIN_COL_INDEX,
                      f'=IF({email_letter}2="","",LOWER(TRIM(MID({email_letter}2,FIND("@",{email_letter}2)+1,100))))'),
    ]

    service.spreadsheets().batchUpdate(
        spreadsheetId=SPREADSHEET_ID,
        body={"requests": requests},
    ).execute()

    # ── Copiar fórmulas hacia abajo ──
    for col in [DATE_COL_INDEX, WEEK_COL_INDEX, MONTH_COL_INDEX, QUARTER_COL_INDEX, DOMAIN_COL_INDEX]:
        service.spreadsheets().batchUpdate(
            spreadsheetId=SPREADSHEET_ID,
            body={"requests": [{"copyPaste": {
                "source": {"sheetId": BASE_SHEET_ID, "startRowIndex": 1, "endRowIndex": 2,
                           "startColumnIndex": col, "endColumnIndex": col + 1},
                "destination": {"sheetId": BASE_SHEET_ID, "startRowIndex": 1, "endRowIndex": last_row,
                                "startColumnIndex": col, "endColumnIndex": col + 1},
                "pasteType": "PASTE_FORMULA",
            }}]},
        ).execute()

    # ── 8 Pivots: Personas + Empresas ──
    _crear_pivot(service, SPREADSHEET_ID, BASE_SHEET_ID, last_row, DOMAIN_COL_INDEX,
                 "Correos Personas Diario", DATE_COL_INDEX, EMAIL_COL_INDEX)
    _crear_pivot(service, SPREADSHEET_ID, BASE_SHEET_ID, last_row, DOMAIN_COL_INDEX,
                 "Correos Personas Semanal", WEEK_COL_INDEX, EMAIL_COL_INDEX)
    _crear_pivot(service, SPREADSHEET_ID, BASE_SHEET_ID, last_row, DOMAIN_COL_INDEX,
                 "Correos Personas Mensual", MONTH_COL_INDEX, EMAIL_COL_INDEX)
    _crear_pivot(service, SPREADSHEET_ID, BASE_SHEET_ID, last_row, DOMAIN_COL_INDEX,
                 "Correos Personas Quarterly", QUARTER_COL_INDEX, EMAIL_COL_INDEX)

    _crear_pivot(service, SPREADSHEET_ID, BASE_SHEET_ID, last_row, DOMAIN_COL_INDEX,
                 "Correos Empresas Diario", DATE_COL_INDEX, DOMAIN_COL_INDEX)
    _crear_pivot(service, SPREADSHEET_ID, BASE_SHEET_ID, last_row, DOMAIN_COL_INDEX,
                 "Correos Empresas Semanal", WEEK_COL_INDEX, DOMAIN_COL_INDEX)
    _crear_pivot(service, SPREADSHEET_ID, BASE_SHEET_ID, last_row, DOMAIN_COL_INDEX,
                 "Correos Empresas Mensual", MONTH_COL_INDEX, DOMAIN_COL_INDEX)
    _crear_pivot(service, SPREADSHEET_ID, BASE_SHEET_ID, last_row, DOMAIN_COL_INDEX,
                 "Correos Empresas Quarterly", QUARTER_COL_INDEX, DOMAIN_COL_INDEX)

    # ── Reply Rates ──
    _process_reply_rates(df_original, spreadsheet)

    return {"rows": len(df), "pivots": 8, "reply_rate_sheets": 4}


def _process_reply_rates(df_original, spreadsheet):
    """Calculate reply rates and upload as separate sheets"""
    df_reply = df_original.copy()

    df_reply["delivery_date"] = pd.to_datetime(
        df_original["Delivery date"],
        format="%a, %d %b %Y %H:%M:%S",
        errors="coerce",
    ).dt.date

    df_reply["Replied"] = pd.to_numeric(df_reply["Replied"], errors="coerce")
    df_reply["Delivered"] = pd.to_numeric(df_reply["Delivered"], errors="coerce")
    df_reply["Contacted"] = pd.to_numeric(df_reply["Contacted"], errors="coerce")

    df_reply["valid_base"] = (df_reply["Delivered"] == 1) & (df_reply["Contacted"] == 1)
    df_reply["valid_reply"] = (df_reply["Replied"] == 1) & (df_reply["Contacted"] == 1)

    df_reply["day"] = df_reply["delivery_date"].astype(str)
    df_reply["week"] = df_reply["delivery_date"].apply(
        lambda x: f"{x.strftime('%V')} - {x.strftime('%y')}" if pd.notnull(x) else None
    )
    df_reply["month"] = df_reply["delivery_date"].apply(
        lambda x: f"{x.strftime('%m')} - {x.strftime('%y')}" if pd.notnull(x) else None
    )
    df_reply["quarter"] = df_reply["delivery_date"].apply(
        lambda x: f"Q{((x.month - 1) // 3) + 1} - {x.strftime('%y')}" if pd.notnull(x) else None
    )

    for nombre, col in [
        ("Reply Rate Diario", "day"),
        ("Reply Rate Semanal", "week"),
        ("Reply Rate Mensual", "month"),
        ("Reply Rate Quarterly", "quarter"),
    ]:
        result_df = _calcular_reply_rate(df_reply, col)
        _subir_reply_sheet(spreadsheet, nombre, result_df)


def _calcular_reply_rate(df, periodo_col):
    resultados = []
    for periodo, grupo in df.groupby(periodo_col):
        base = grupo.loc[grupo["valid_base"], "Contact email"].nunique()
        replies = grupo.loc[grupo["valid_reply"], "Contact email"].nunique()
        reply_rate = round(replies / base, 4) if base > 0 else 0

        if hasattr(periodo, "strftime"):
            periodo = periodo.strftime("%Y-%m-%d")

        resultados.append({
            "Period": str(periodo),
            "Base": int(base),
            "Replies": int(replies),
            "Reply Rate": reply_rate,
        })

    return pd.DataFrame(resultados).sort_values("Period")


def _subir_reply_sheet(spreadsheet, nombre, dataframe):
    new_sheet = spreadsheet.add_worksheet(
        title=nombre,
        rows=str(len(dataframe) + 10),
        cols="5",
    )
    clean_df = dataframe.replace([float("inf"), float("-inf")], "").fillna("")
    new_sheet.update([clean_df.columns.tolist()] + clean_df.values.tolist())


# ── Helpers (same as carga_personas) ──

def _header_cell(sheet_id, col_index, value):
    return {"updateCells": {
        "start": {"sheetId": sheet_id, "rowIndex": 0, "columnIndex": col_index},
        "rows": [{"values": [{"userEnteredValue": {"stringValue": value}}]}],
        "fields": "userEnteredValue",
    }}


def _formula_cell(sheet_id, col_index, formula):
    return {"updateCells": {
        "start": {"sheetId": sheet_id, "rowIndex": 1, "columnIndex": col_index},
        "rows": [{"values": [{"userEnteredValue": {"formulaValue": formula}}]}],
        "fields": "userEnteredValue",
    }}


def _crear_pivot(service, spreadsheet_id, base_sheet_id, last_row, last_col, nombre, fila_index, value_index):
    res = service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": [{"addSheet": {"properties": {"title": nombre}}}]},
    ).execute()

    pid = res["replies"][0]["addSheet"]["properties"]["sheetId"]

    pivot = {"pivotTable": {
        "source": {
            "sheetId": base_sheet_id,
            "startRowIndex": 0, "startColumnIndex": 0,
            "endRowIndex": last_row, "endColumnIndex": last_col + 1,
        },
        "rows": [{"sourceColumnOffset": fila_index, "showTotals": True, "sortOrder": "ASCENDING"}],
        "values": [{"sourceColumnOffset": value_index, "summarizeFunction": "COUNTUNIQUE"}],
    }}

    service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": [{"updateCells": {
            "start": {"sheetId": pid, "rowIndex": 0, "columnIndex": 0},
            "rows": [{"values": [pivot]}],
            "fields": "pivotTable",
        }}]},
    ).execute()
