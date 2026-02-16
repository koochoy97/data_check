"""
Bloque 1: Carga de Personas + Empresas
Adaptado de auditoría_elt.py (líneas 19-277)

Input: CSV con columnas "Added On", "Email"
Output: Agrega a un spreadsheet existente:
  - Hoja "Base Personas" con datos + columnas derivadas
  - 8 pivots: Personas/Empresas × Día/Semana/Mes/Quarter
"""
import pandas as pd
from openpyxl.utils import get_column_letter


def procesar_carga(csv_path, spreadsheet, service):
    """
    Args:
        csv_path: Path al CSV de People (Basic fields)
        spreadsheet: gspread Spreadsheet object (ya creado)
        service: Google Sheets API service
    Returns:
        dict con metadata del procesamiento
    """
    SPREADSHEET_ID = spreadsheet.id

    # ── Leer CSV ──
    df = pd.read_csv(csv_path)

    # ── Columna derivada: Added On Date ──
    df["Added On Date"] = pd.to_datetime(
        df["Added On"],
        format="%m/%d/%Y %I:%M %p",
        errors="coerce",
    ).dt.strftime("%d/%m/%Y")

    df = df.replace([float("inf"), float("-inf")], "")
    df = df.fillna("")
    df = df.astype(str)

    # ── Crear hoja "Base Personas" ──
    sheet = spreadsheet.add_worksheet(
        title="Base Personas",
        rows=str(len(df) + 10),
        cols=str(len(df.columns) + 10),
    )
    sheet.update([df.columns.tolist()] + df.values.tolist())

    BASE_SHEET_ID = sheet._properties["sheetId"]
    last_row = len(df) + 1

    # ── Índices de columnas ──
    EMAIL_COL_INDEX = df.columns.get_loc("Email")
    EMAIL_LETTER = get_column_letter(EMAIL_COL_INDEX + 1)

    DATE_REAL_INDEX = len(df.columns)
    WEEK_INDEX = DATE_REAL_INDEX + 1
    MONTH_INDEX = DATE_REAL_INDEX + 2
    QUARTER_INDEX = DATE_REAL_INDEX + 3
    DOMAIN_INDEX = DATE_REAL_INDEX + 4

    # ── Expandir grid ──
    service.spreadsheets().batchUpdate(
        spreadsheetId=SPREADSHEET_ID,
        body={"requests": [{"updateSheetProperties": {
            "properties": {"sheetId": BASE_SHEET_ID, "gridProperties": {"columnCount": DOMAIN_INDEX + 1}},
            "fields": "gridProperties.columnCount",
        }}]},
    ).execute()

    # ── Columnas derivadas con fórmulas ──
    date_letter = get_column_letter(df.columns.get_loc("Added On Date") + 1)
    requests = [
        # Fecha real (Date type)
        _header_cell(BASE_SHEET_ID, DATE_REAL_INDEX, "Added On Date (Date)"),
        _formula_cell(BASE_SHEET_ID, DATE_REAL_INDEX,
                      f'=IF({date_letter}2="","",DATE(RIGHT({date_letter}2,4),MID({date_letter}2,4,2),LEFT({date_letter}2,2)))'),
        # Semana
        _header_cell(BASE_SHEET_ID, WEEK_INDEX, "Semana del año"),
        _formula_cell(BASE_SHEET_ID, WEEK_INDEX,
                      f'=IF({get_column_letter(DATE_REAL_INDEX+1)}2="","",TEXT(WEEKNUM({get_column_letter(DATE_REAL_INDEX+1)}2,21),"00")&" - "&TEXT({get_column_letter(DATE_REAL_INDEX+1)}2,"YY"))'),
        # Mes
        _header_cell(BASE_SHEET_ID, MONTH_INDEX, "Mes - Año"),
        _formula_cell(BASE_SHEET_ID, MONTH_INDEX,
                      f'=IF({get_column_letter(DATE_REAL_INDEX+1)}2="","",TEXT({get_column_letter(DATE_REAL_INDEX+1)}2,"MM")&" - "&TEXT({get_column_letter(DATE_REAL_INDEX+1)}2,"YY"))'),
        # Quarter
        _header_cell(BASE_SHEET_ID, QUARTER_INDEX, "Quarter - Año"),
        _formula_cell(BASE_SHEET_ID, QUARTER_INDEX,
                      f'=IF({get_column_letter(DATE_REAL_INDEX+1)}2="","","Q"&ROUNDUP(MONTH({get_column_letter(DATE_REAL_INDEX+1)}2)/3,0)&" - "&TEXT({get_column_letter(DATE_REAL_INDEX+1)}2,"YY"))'),
        # Email Domain
        _header_cell(BASE_SHEET_ID, DOMAIN_INDEX, "Email Domain"),
        _formula_cell(BASE_SHEET_ID, DOMAIN_INDEX,
                      f'=IF({EMAIL_LETTER}2="","",LOWER(TRIM(MID({EMAIL_LETTER}2,FIND("@",{EMAIL_LETTER}2)+1,100))))'),
    ]

    service.spreadsheets().batchUpdate(
        spreadsheetId=SPREADSHEET_ID,
        body={"requests": requests},
    ).execute()

    # ── Copiar fórmulas hacia abajo ──
    for col in [DATE_REAL_INDEX, WEEK_INDEX, MONTH_INDEX, QUARTER_INDEX, DOMAIN_INDEX]:
        service.spreadsheets().batchUpdate(
            spreadsheetId=SPREADSHEET_ID,
            body={"requests": [{"copyPaste": {
                "source": {"sheetId": BASE_SHEET_ID, "startRowIndex": 1, "endRowIndex": 2,
                           "startColumnIndex": col, "endColumnIndex": col + 1},
                "destination": {"sheetId": BASE_SHEET_ID, "startRowIndex": 1, "endRowIndex": last_row,
                                "startColumnIndex": col, "endColumnIndex": col + 1},
                "pasteType": "PASTE_FORMULA",
            }}]},
        ).execute()

    # ── Crear 8 pivots ──
    _crear_pivot(service, SPREADSHEET_ID, BASE_SHEET_ID, last_row, DOMAIN_INDEX,
                 "Pivot Carga Diaria", DATE_REAL_INDEX, EMAIL_COL_INDEX)
    _crear_pivot(service, SPREADSHEET_ID, BASE_SHEET_ID, last_row, DOMAIN_INDEX,
                 "Pivot Semanal", WEEK_INDEX, EMAIL_COL_INDEX)
    _crear_pivot(service, SPREADSHEET_ID, BASE_SHEET_ID, last_row, DOMAIN_INDEX,
                 "Pivot Mensual", MONTH_INDEX, EMAIL_COL_INDEX)
    _crear_pivot(service, SPREADSHEET_ID, BASE_SHEET_ID, last_row, DOMAIN_INDEX,
                 "Pivot Quarterly", QUARTER_INDEX, EMAIL_COL_INDEX)

    _crear_pivot(service, SPREADSHEET_ID, BASE_SHEET_ID, last_row, DOMAIN_INDEX,
                 "Empresas Diario", DATE_REAL_INDEX, DOMAIN_INDEX)
    _crear_pivot(service, SPREADSHEET_ID, BASE_SHEET_ID, last_row, DOMAIN_INDEX,
                 "Empresas Semanal", WEEK_INDEX, DOMAIN_INDEX)
    _crear_pivot(service, SPREADSHEET_ID, BASE_SHEET_ID, last_row, DOMAIN_INDEX,
                 "Empresas Mensual", MONTH_INDEX, DOMAIN_INDEX)
    _crear_pivot(service, SPREADSHEET_ID, BASE_SHEET_ID, last_row, DOMAIN_INDEX,
                 "Empresas Quarterly", QUARTER_INDEX, DOMAIN_INDEX)

    return {"rows": len(df), "pivots": 8}


# ── Helpers ──

def _header_cell(sheet_id, col_index, value):
    return {"updateCells": {
        "start": {"sheetId": sheet_id, "rowIndex": 0, "columnIndex": col_index},
        "rows": [{"values": [{"userEnteredValue": {"stringValue": value}}]}],
        "fields": "userEnteredValue",
    }}


def _formula_cell(sheet_id, col_index, formula):
    return {"updateCells": {
        "start": {"sheetId": sheet_id, "rowIndex": 1, "columnIndex": col_index},
        "rows": [{"values": [{"userEnteredValue": {"formulaValue": formula}}]}],
        "fields": "userEnteredValue",
    }}


def _crear_pivot(service, spreadsheet_id, base_sheet_id, last_row, last_col, nombre, fila_index, value_index):
    res = service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": [{"addSheet": {"properties": {"title": nombre}}}]},
    ).execute()

    pid = res["replies"][0]["addSheet"]["properties"]["sheetId"]

    pivot = {"pivotTable": {
        "source": {
            "sheetId": base_sheet_id,
            "startRowIndex": 0, "startColumnIndex": 0,
            "endRowIndex": last_row, "endColumnIndex": last_col + 1,
        },
        "rows": [{"sourceColumnOffset": fila_index, "showTotals": True, "sortOrder": "ASCENDING"}],
        "values": [{"sourceColumnOffset": value_index, "summarizeFunction": "COUNTUNIQUE"}],
    }}

    service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": [{"updateCells": {
            "start": {"sheetId": pid, "rowIndex": 0, "columnIndex": 0},
            "rows": [{"values": [pivot]}],
            "fields": "pivotTable",
        }}]},
    ).execute()
